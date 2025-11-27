import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class MonitorPromocoesComPrecosExcel:
    def __init__(self, access_token, user_id, max_threads=10, output_file="relatorio_promos.xlsx"):
        self.access_token = access_token
        self.user_id = user_id
        self.headers = {'Authorization': f'Bearer {access_token}'}
        self.max_threads = max_threads
        self.output_file = output_file

    # -------------------------
    # BUSCAR PROMOÇÕES
    # -------------------------
    def get_promocoes(self):
        url = f"https://api.mercadolibre.com/seller-promotions/users/{self.user_id}?app_version=v2"
        resp = requests.get(url, headers=self.headers)
        if resp.status_code != 200:
            print("❌ Erro ao buscar promoções:", resp.status_code)
            return []
        return resp.json().get("results", [])

    # -------------------------
    # BUSCAR ITENS DE CADA PROMOÇÃO
    # -------------------------
    def get_itens_promocao(self, promo_id, promo_type):
        items = []
        url = f"https://api.mercadolibre.com/seller-promotions/promotions/{promo_id}/items"
        search_after = None

        while True:
            params = {
                "promotion_type": promo_type,
                "app_version": "v2",
                "limit": 50
            }
            if search_after:
                params["search_after"] = search_after

            resp = requests.get(url, headers=self.headers, params=params)
            if resp.status_code != 200:
                print("❌ Erro ao buscar itens da promoção", promo_id, resp.status_code)
                break

            data = resp.json()
            items.extend(data.get("results", []))

            paging = data.get("paging", {})
            search_after = paging.get("searchAfter")
            if not search_after:
                break

        return items

    # -------------------------
    # BUSCAR PREÇO ATUAL
    # -------------------------
    def get_preco_item(self, item_id):
        url = f"https://api.mercadolibre.com/items/{item_id}/sale_price?context=channel_marketplace,buyer_loyalty_3"

        try:
            resp = requests.get(url, headers=self.headers, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                return data.get("amount", None), data.get("currency_id", None)
        except Exception as e:
            print("❌ Erro ao buscar preço do item", item_id, e)

        return None, None

    # -------------------------
    # GERAR DATAFRAME COMPLETO
    # -------------------------
    def gerar_dados(self):
        promos = self.get_promocoes()
        todos_itens = {}

        for promo in promos:
            pid = promo.get("id")
            ptype = promo.get("type")
            pname = promo.get("name", "")

            itens = self.get_itens_promocao(pid, ptype)
            print(f"📦 Promoção {pid} — {len(itens)} itens")

            for it in itens:
                iid = it.get("id")
                if not iid:
                    continue

                if iid not in todos_itens:
                    todos_itens[iid] = []

                todos_itens[iid].append({
                    "promotion_id": pid,
                    "promotion_type": ptype,
                    "promotion_name": pname,
                    "status": it.get("status"),
                    "price": it.get("price"),
                    "original_price": it.get("original_price"),
                    "min_discounted_price": it.get("min_discounted_price"),
                    "max_discounted_price": it.get("max_discounted_price"),
                    "suggested_discounted_price": it.get("suggested_discounted_price"),
                })

        linhas = []
        total = len(todos_itens)
        count = 0

        with ThreadPoolExecutor(max_workers=self.max_threads) as executor:
            futures = {executor.submit(self.get_preco_item, iid): iid for iid in todos_itens}

            for fut in as_completed(futures):
                iid = futures[fut]
                count += 1
                preco, moeda = fut.result()

                print(f"[{count}/{total}] 💰 Preço item {iid} → {preco}")

                for pinfo in todos_itens[iid]:
                    linha = {
                        "item_id": iid,
                        "moeda": moeda,
                        "preco_mercado": preco,
                        **pinfo
                    }
                    linhas.append(linha)

        df = pd.DataFrame(linhas)
        return df

    # -------------------------
    # SALVAR EXCEL + ALERTAS
    # -------------------------
    def salvar_excel_com_alertas(self, df):
        wb = Workbook()

        # ABA 1 — Dados completos
        ws_data = wb.active
        ws_data.title = "Todos os Dados"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)

        # Mapeia preço base da promoção (status=started)
        price_started = {}

        for _, row in df.iterrows():
            if row["status"] == "started":
                iid = row["item_id"]
                val = row.get("price")
                if iid not in price_started or val < price_started[iid]:
                    price_started[iid] = val

        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

        # Pintar alertas
        for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
            iid = row[0].value  # item_id
            status = row[6].value  # coluna status
            max_disc = row[10].value  # coluna max_discounted_price

            base_price = price_started.get(iid)

            try:
                if status == "candidate" and base_price and max_disc and float(max_disc) > float(base_price):
                    for cell in row:
                        cell.fill = red_fill
            except:
                pass

        # ABA 2 — Somente alertas filtrados
        ws_alert = wb.create_sheet("Alertas")

        df_alertas = df.groupby("item_id").apply(
            lambda g: g[
                (g["status"] == "candidate")
                & (g["max_discounted_price"] > g.loc[g["status"] == "started", "price"].min(skipna=True))
            ]
        ).reset_index(drop=True)

        for r in dataframe_to_rows(df_alertas, index=False, header=True):
            ws_alert.append(r)

        wb.save(self.output_file)
        print("📁 Excel salvo em", self.output_file)

    # -------------------------
    # GERAR JSON PARA O SITE
    # -------------------------
    def salvar_json(self, df, caminho="site/dados.json"):
        df.to_json(caminho, orient="records", force_ascii=False, indent=2)
        print(f"🌐 JSON salvo em {caminho}")

    # -------------------------
    # EXECUTAR TUDO
    # -------------------------
    def run(self):
        df = self.gerar_dados()
        self.salvar_excel_com_alertas(df)
        self.salvar_json(df)


# -----------------------------------
# EXECUÇÃO PRINCIPAL
# -----------------------------------
if __name__ == "__main__":
    ACCESS = "APP_USR-5767896809769647-112709-072905d05838254b5df442045e20d84c-415176739"
    USER = "415176739"

    monitor = MonitorPromocoesComPrecosExcel(
        ACCESS,
        USER,
        max_threads=20,
        output_file="relatorio.xlsx"
    )

    monitor.run()

